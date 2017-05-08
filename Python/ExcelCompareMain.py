#!/usr/bin/env python

from openpyxl import load_workbook
import sys

__author__ = "Pascal van de Wijdeven"
__copyright__ = "Copyright 2016"
__credits__ = ["Pascal van de Wijdeven"]
__license__ = "GPL"
__version__ = "0.0.1"
__maintainer__ = "Pascal van de Wijdeven"
__email__ = "nospam_pvdwijdeven@gmail.dot.com"
__status__ = "Initial"

###############################################################################
# Both input_old and input_new files shall have headers in the first row,
# starting in column A and have no empty headers before last column. Every
# header shall be unique. Headers in old/new can be mixed, old format will be
# maintained. New headers and deleted headers will be marked up.
#
# Only first sheet will be compared (for now)
###############################################################################

WORK_PATH = "C:\\Compare\\"
INPUT_OLD = "old.xlsx"
INPUT_NEW = "new.xlsx"
OUTPUT_COMPARED = "difference.xlsx"

if len(sys.argv) >= 2:
    INPUT_OLD = (str(sys.argv[1]))
if len(sys.argv) >= 3:
    INPUT_NEW = (str(sys.argv[2]))
if len(sys.argv) >= 4:
    OUTPUT_COMPARED = (str(sys.argv[3]))

input_old = load_workbook(WORK_PATH + INPUT_OLD)
input_new = load_workbook(WORK_PATH + INPUT_NEW)

for ws in [input_old.active, input_new.active]:
    header = []

    for x in range(1, 1000, 1):
        cur_val = ws.cell(row=1, column=x).value
        if cur_val is not None:
            header.append(cur_val)
            lastColumn = x
        else:
            break
    if ws == input_old.active:
        header_old = header
    else:
        header_new = header
print(header_old)
print(header_new)

